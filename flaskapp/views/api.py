from flask import Blueprint, request, jsonify, send_file, make_response, Response
from ..models import db, Tender
from .. import logger
from .main import login_required
from ..utils import (
    allowed_file, 
    get_friendly_error_message, 
    find_duplicate,
    get_duplicate_type,
    format_amount
)
import pandas as pd
from datetime import datetime
import openpyxl
from io import BytesIO
from urllib.parse import quote  # 新增引用
from werkzeug.datastructures import MultiDict

api_bp = Blueprint('api', __name__)

def build_filter_query(args):
    """构建数据库查询，添加筛选条件"""
    query = Tender.query
    
    # 项目名称搜索（支持多关键词）
    project_name = args.get('project_name')
    if project_name:
        keywords = [k.strip() for k in project_name.split() if k.strip()]
        for keyword in keywords:
            query = query.filter(Tender.project_name.ilike(f'%{keyword}%'))
    
    # 发布时间范围
    publish_time_start = args.get('publish_time_start')
    publish_time_end = args.get('publish_time_end')
    if publish_time_start:
        query = query.filter(Tender.publish_time >= publish_time_start)
    if publish_time_end:
        query = query.filter(Tender.publish_time <= f"{publish_time_end} 23:59:59")
    
    # 投标截止时间范围
    bid_deadline_start = args.get('bid_deadline_start')
    bid_deadline_end = args.get('bid_deadline_end')
    if bid_deadline_start:
        query = query.filter(Tender.bid_deadline >= bid_deadline_start)
    if bid_deadline_end:
        query = query.filter(Tender.bid_deadline <= f"{bid_deadline_end} 23:59:59")
    
    # 招标阶段
    bid_stage = args.get('bid_stage')
    if bid_stage:
        query = query.filter(Tender.bid_stage == bid_stage)
    
    # 项目类型
    project_type = args.get('project_type')
    if project_type:
        query = query.filter(Tender.project_type == project_type)
    
    # 契合度
    matching_degree = args.get('matching_degree')
    if matching_degree:
        query = query.filter(Tender.matching_degree == matching_degree)
    
    # 地区筛选
    province = args.get('province')
    city = args.get('city')
    if province:
        query = query.filter(Tender.province == province)
    if city:
        query = query.filter(Tender.city == city)
    
    # 招标单位
    bidding_unit = args.get('bidding_unit')
    if bidding_unit:
        query = query.filter(Tender.bidding_unit.ilike(f'%{bidding_unit}%'))
    
    # 金额范围
    bid_amount_min = args.get('bid_amount_min', type=float)
    bid_amount_max = args.get('bid_amount_max', type=float)
    if bid_amount_min is not None:
        query = query.filter(Tender.bid_amount >= bid_amount_min)
    if bid_amount_max is not None:
        query = query.filter(Tender.bid_amount <= bid_amount_max)
    
    return query

@api_bp.route('/bids')
def get_bids():
    """获取招标数据列表"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = 20
        
        query = build_filter_query(request.args)
        query = query.order_by(Tender.publish_time.desc())
        pagination = query.paginate(page=page, per_page=per_page)
        
        items = [item.to_dict() for item in pagination.items]
        
        return jsonify({
            'items': items,
            'total': pagination.total,
            'pages': pagination.pages,
            'current_page': page
        })
        
    except Exception as e:
        logger.error(f"Error in get_bids: {str(e)}")
        return jsonify({'error': str(e)}), 500

@api_bp.route('/import', methods=['POST'])
@login_required
def import_data():
    """导入数据"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': '未选择文件'}), 400
        
        file = request.files['file']
        if not file.filename:
            return jsonify({'error': '文件名为空'}), 400
            
        if not allowed_file(file.filename):
            return jsonify({'error': '不支持的文件类型'}), 400
            
        try:
            df = pd.read_excel(file)
            logger.info(f'开始导入Excel文件，共{len(df)}行数据')
            
            # 列名映射
            column_mapping = {
                '信息录入时间': 'entry_time',
                '项目名称': 'project_name',
                '关键词': 'keywords',
                '契合度': 'matching_degree',
                '项目类型': 'project_type',
                '项目编号': 'project_number',
                '发布省份': 'province',
                '发布市级': 'city',
                '发布区级': 'district',
                '信息发布时间': 'publish_time',
                '招标阶段': 'bid_stage',
                '报名截止时间': 'registration_deadline',
                '投标截止时间': 'bid_deadline',
                '招标金额（元）': 'bid_amount',
                '招标单位': 'bidding_unit',
                '招标单位.1': 'bidding_unit_extra1',
                '招标单位.2': 'bidding_unit_extra2',
                '代理单位': 'agency_unit',
                '代理单位.1': 'agency_unit_extra1',
                '代理单位.2': 'agency_unit_extra2',
                '官网查看地址': 'website_url'
            }
            
            # 重命名列
            df = df.rename(columns=column_mapping)
            
            success_count = 0
            duplicate_count = 0
            error_count = 0
            error_details = []
            duplicate_records = []

            for index, row in df.iterrows():
                try:
                    tender = Tender(
                        entry_time=pd.to_datetime(row.get('entry_time')) if pd.notnull(row.get('entry_time')) else datetime.now(),
                        project_name=str(row.get('project_name')).strip() if pd.notnull(row.get('project_name')) else None,
                        keywords=str(row.get('keywords')).strip() if pd.notnull(row.get('keywords')) else None,
                        project_number=str(row.get('project_number')).strip() if pd.notnull(row.get('project_number')) else None,
                        province=str(row.get('province')).strip() if pd.notnull(row.get('province')) else None,
                        city=str(row.get('city')).strip() if pd.notnull(row.get('city')) else None,
                        district=str(row.get('district')).strip() if pd.notnull(row.get('district')) else None,
                        publish_time=pd.to_datetime(row.get('publish_time')) if pd.notnull(row.get('publish_time')) else None,
                        bid_stage=str(row.get('bid_stage')).strip() if pd.notnull(row.get('bid_stage')) else None,
                        registration_deadline=pd.to_datetime(row.get('registration_deadline')) if pd.notnull(row.get('registration_deadline')) else None,
                        bid_deadline=pd.to_datetime(row.get('bid_deadline')) if pd.notnull(row.get('bid_deadline')) else None,
                        bid_amount=float(row.get('bid_amount')) if pd.notnull(row.get('bid_amount')) else None,
                        bidding_unit=str(row.get('bidding_unit')).strip() if pd.notnull(row.get('bidding_unit')) else None,
                        bidding_unit_extra1=str(row.get('bidding_unit_extra1')).strip() if pd.notnull(row.get('bidding_unit_extra1')) else None,
                        bidding_unit_extra2=str(row.get('bidding_unit_extra2')).strip() if pd.notnull(row.get('bidding_unit_extra2')) else None,
                        agency_unit=str(row.get('agency_unit')).strip() if pd.notnull(row.get('agency_unit')) else None,
                        agency_unit_extra1=str(row.get('agency_unit_extra1')).strip() if pd.notnull(row.get('agency_unit_extra1')) else None,
                        agency_unit_extra2=str(row.get('agency_unit_extra2')).strip() if pd.notnull(row.get('agency_unit_extra2')) else None,
                        website_url=str(row.get('website_url')).strip() if pd.notnull(row.get('website_url')) else None,
                        project_type=str(row.get('project_type')).strip() if pd.notnull(row.get('project_type')) else None,
                        matching_degree=str(row.get('matching_degree')).strip() if pd.notnull(row.get('matching_degree')) else None
                    )

                    # 检查重复
                    existing = find_duplicate(tender)
                    if existing:
                        duplicate_count += 1
                        # 添加重复记录的详细对比信息
                        duplicate_info = {
                            'row_number': index + 2,
                            'new_record': {
                                'project_name': tender.project_name,
                                'project_number': tender.project_number,
                                'publish_time': tender.publish_time.strftime('%Y-%m-%d') if tender.publish_time else None,
                                'bid_stage': tender.bid_stage,
                                'bidding_unit': tender.bidding_unit,
                                'bid_deadline': tender.bid_deadline.strftime('%Y-%m-%d') if tender.bid_deadline else None,
                                'bid_amount': format_amount(tender.bid_amount) if tender.bid_amount else None,
                                'agency_unit': tender.agency_unit
                            },
                            'existing_record': {
                                'project_name': existing.project_name,
                                'project_number': existing.project_number,
                                'publish_time': existing.publish_time.strftime('%Y-%m-%d') if existing.publish_time else None,
                                'bid_stage': existing.bid_stage,
                                'bidding_unit': existing.bidding_unit,
                                'bid_deadline': existing.bid_deadline.strftime('%Y-%m-%d') if existing.bid_deadline else None,
                                'bid_amount': format_amount(existing.bid_amount) if existing.bid_amount else None,
                                'agency_unit': existing.agency_unit
                            },
                            'duplicate_type': get_duplicate_type(tender, existing)
                        }
                        duplicate_records.append(duplicate_info)
                        continue

                    db.session.add(tender)
                    success_count += 1

                except Exception as e:
                    error_count += 1
                    error_details.append({
                        'row_number': index + 2,
                        'error_message': get_friendly_error_message(e),
                        'data': {
                            'project_name': row.get('project_name', '--'),
                            'project_number': row.get('project_number', '--'),
                            'publish_time': row.get('publish_time', '--'),
                            'bid_amount': row.get('bid_amount', '--')
                        }
                    })
                    logger.error(f'处理第{index+2}行数据时出错: {str(e)}')
                    continue
            
            db.session.commit()
            logger.info(f'Excel导入完成：成功{success_count}条，重复{duplicate_count}条，失败{error_count}条')
            
            return jsonify({
                'success': True,
                'total': len(df),
                'success_count': success_count,
                'duplicate_count': duplicate_count,
                'error_count': error_count,
                'error_details': error_details,
                'duplicate_records': duplicate_records,
                'message': f'导入完成：成功{success_count}条，重复{duplicate_count}条，失败{error_count}条'
            })
            
        except Exception as e:
            db.session.rollback()
            logger.error(f'导入过程出错: {str(e)}')
            return jsonify({
                'error': get_friendly_error_message(e),
                'success': False
            }), 500
            
    except Exception as e:
        logger.error(f'导入过程出错: {str(e)}')
        return jsonify({
            'error': get_friendly_error_message(e),
            'success': False
        }), 500

@api_bp.route('/export', methods=['POST'])
def export_data():
    """导出数据"""
    try:
        # 获取导出类型和数据
        export_type = request.json.get('exportType', 'all')  # 'selected' 或 'all'
        selected_data = request.json.get('selectedData', [])
        filter_params = request.json.get('filterParams', {})
        
        # 根据导出类型获取数据
        if export_type == 'selected' and selected_data:
            # 使用选中的数据
            data = selected_data
        else:
            # 使用筛选条件获取所有数据
            # 将 filter_params 转换为 MultiDict 格式
            filter_args = MultiDict()
            for key, value in filter_params.items():
                if value:  # 只添加非空值
                    filter_args.add(key, value)
            
            query = build_filter_query(filter_args)
            data = [{
                'project_name': bid.project_name,
                'publish_time': bid.publish_time.strftime('%Y-%m-%d') if bid.publish_time else '未设置',
                'bid_deadline': bid.bid_deadline.strftime('%Y-%m-%d') if bid.bid_deadline else '未设置',
                'bid_stage': bid.bid_stage or '未设置',
                'bid_amount': format_amount(bid.bid_amount),
                'bidding_unit': bid.bidding_unit or '未设置',
                'project_type': bid.project_type or '未设置',
                'matching_degree': bid.matching_degree or '未设置',
                'province': bid.province or '未设置',
                'city': bid.city or '未设置',
                'website_url': bid.website_url or '未设置'  # 添加链接字段
            } for bid in query.all()]
        
        if not data:
            return jsonify({'error': '没有数据可导出'}), 400

        # 创建Excel文件
        output = BytesIO()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = '招标数据'
        
        # 设置列名
        headers = {
            'project_name': '项目名称',
            'publish_time': '发布时间',
            'bid_deadline': '投标截止',
            'bid_stage': '招标阶段',
            'bid_amount': '招标金额',
            'bidding_unit': '招标单位',
            'project_type': '项目类型',
            'matching_degree': '契合度',
            'province': '省份',
            'city': '城市',
            'website_url': '原文链接'  # 添加链接列名
        }
        
        # 设置表头样式
        header_style = openpyxl.styles.NamedStyle(name='header')
        header_style.font = openpyxl.styles.Font(bold=True)
        header_style.fill = openpyxl.styles.PatternFill(start_color='f2f2f2', end_color='f2f2f2', fill_type='solid')
        header_style.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        
        # 写入表头
        for col, header in enumerate(headers.values(), 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.style = header_style
        
        # 设置单元格样式
        date_style = openpyxl.styles.NamedStyle(name='date')
        date_style.alignment = openpyxl.styles.Alignment(horizontal='center')
        
        amount_style = openpyxl.styles.NamedStyle(name='amount')
        amount_style.alignment = openpyxl.styles.Alignment(horizontal='right')
        
        # 写入数据
        for row, item in enumerate(data, 2):
            for col, field in enumerate(headers.keys(), 1):
                cell = worksheet.cell(row=row, column=col, value=item.get(field, '未设置'))
                if field in ['publish_time', 'bid_deadline']:
                    cell.style = date_style
                elif field == 'bid_amount':
                    cell.style = amount_style
                elif field == 'website_url' and item.get(field) != '未设置':
                    # 为链接添加超链接
                    cell.hyperlink = item.get(field)
                    cell.font = openpyxl.styles.Font(color='0000FF', underline='single')
        
        # 调整列宽
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column].width = adjusted_width
        
        # 保存到内存
        workbook.save(output)
        output.seek(0)
        
        # 生成带精确时间的文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'招标_{timestamp}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
            conditional=True
        )
        
    except Exception as e:
        logger.error(f'导出数据时发生错误: {str(e)}')
        return jsonify({'error': f'导出失败: {str(e)}'}), 500

@api_bp.route('/options')
def get_options():
    """获取选项数据"""
    try:
        # 获取所有不重复的选项
        bid_stages = db.session.query(Tender.bid_stage).distinct().filter(Tender.bid_stage.isnot(None)).all()
        project_types = db.session.query(Tender.project_type).distinct().filter(Tender.project_type.isnot(None)).all()
        matching_degrees = ['高', '中', '低']  # 固定的契合度选项
        
        # 获取地区信息
        regions = db.session.query(
            Tender.province,
            Tender.city
        ).distinct().filter(
            Tender.province.isnot(None)
        ).order_by(
            Tender.province,
            Tender.city
        ).all()
        
        # 组织省市数据
        provinces = {}
        for province, city in regions:
            if province:
                if province not in provinces:
                    provinces[province] = set()
                if city:
                    provinces[province].add(city)
        
        # 转换为前端需要的格式
        formatted_regions = [
            {
                'name': province,
                'cities': sorted(list(cities))
            } for province, cities in provinces.items()
        ]
        
        return jsonify({
            'bid_stages': [stage[0] for stage in bid_stages if stage[0]],
            'project_types': [type[0] for type in project_types if type[0]],
            'matching_degrees': matching_degrees,
            'provinces': sorted(formatted_regions, key=lambda x: x['name'])
        })
        
    except Exception as e:
        logger.error(f'获取选项数据时发生错误: {str(e)}')
        return jsonify({'error': str(e)}), 500

@api_bp.route('/bids/count', methods=['GET'])
def get_bids_count():
    """获取符合条件的数据总数"""
    try:
        query = build_filter_query(request.args)
        total = query.count()
        return jsonify({'total': total})
    except Exception as e:
        logger.error(f'获取数据总数时发生错误: {str(e)}')
        return jsonify({'error': str(e)}), 500

@api_bp.route('/bids/<int:bid_id>', methods=['DELETE'])
@login_required
def delete_bid(bid_id):
    """删除单条记录"""
    try:
        tender = Tender.query.get_or_404(bid_id)
        db.session.delete(tender)
        db.session.commit()
        logger.info(f'成功删除记录 ID: {bid_id}')
        return jsonify({'message': '删除成功'})
    except Exception as e:
        db.session.rollback()
        logger.error(f'删除记录时发生错误: {str(e)}')
        return jsonify({'error': '删除失败'}), 500

@api_bp.route('/bids/batch', methods=['DELETE'])
@login_required
def batch_delete_bids():
    """批量删除记录"""
    try:
        bid_ids = request.json.get('ids', [])
        if not bid_ids:
            return jsonify({'error': '未选择要删除的记录'}), 400
            
        # 查找所有要删除的记录
        bids = Tender.query.filter(Tender.id.in_(bid_ids)).all()
        deleted_count = len(bids)
        
        # 执行删除
        for bid in bids:
            db.session.delete(bid)
            
        db.session.commit()
        logger.info(f'成功批量删除 {deleted_count} 条记录')
        
        return jsonify({
            'message': f'成功删除 {deleted_count} 条记录',
            'deleted_count': deleted_count
        })
    except Exception as e:
        db.session.rollback()
        logger.error(f'批量删除记录时发生错误: {str(e)}')
        return jsonify({'error': '删除失败'}), 500

@api_bp.route('/bids/<int:bid_id>', methods=['GET'])
def get_bid_details(bid_id):
    """获取招标详情"""
    try:
        tender = Tender.query.get_or_404(bid_id)
        return jsonify(tender.to_dict())
    except Exception as e:
        logger.error(f'获取招标详情时发生错误: {str(e)}')
        return jsonify({'error': str(e)}), 500

# 其他API路由... 